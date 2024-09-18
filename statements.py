from export_pdf import export_to_pdf
import pandas as pd
from openpyxl import load_workbook
from VARIABLES import comp_month, tms, rms, am_comp_file, csr_comp_file, rm_comp_file


def export_to_excel(excel_file: str, tab: str, dataframe: pd.DataFrame):
    """Writes a pandas dataframe pulled from database.py to the specified Excel file and tab."""
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=tab, index=False)


def am_statement(payees, **kwargs):
    """Exports all comp details to COMP_STATEMENT.xlsx, which can then be used to generate an official statement.
    Optionally, you can pass in an email kwarg to view info for a single rep. Set export=True to run a VBA script to
    generate a PDF statement."""
    email = kwargs.get('email', None)
    print("Generating AM Statements...")
    for am in payees.am_info:
        if email is None or payees.am_info[am]['EMAIL'] in email:
            # get the info for only the current loop rep
            name = am
            eid = payees.am_info[am]['EMAIL']
            rm = payees.am_info[am]['RM_EMAIL']
            terr = payees.am_info[am]['TERR_NM']
            payout_df = payees.tblpayout[payees.tblpayout['EID'] == eid]
            comp_detail_df = payees.am_comp_detail[payees.am_comp_detail['SALES_CREDIT_REP_EMAIL'] == eid]

            excel_file = am_comp_file

            # export the rep's tblPayout info and comp detail info to different tabs on COMP_STATEMENT.xlsx
            export_to_excel(excel_file, 'payout', payout_df)
            export_to_excel(excel_file, 'detail', comp_detail_df)

            # export the rep's name, email, rm email, and territory name to the 'info' tab

            wb = load_workbook(excel_file)
            sheet_name = 'info'
            sheet = wb[sheet_name]
            sheet['B1'].value = name
            sheet['B2'].value = eid
            sheet['B3'].value = rm
            sheet['B4'].value = terr
            if payees.am_info[am]['EMAIL'] in tms:
                sheet['B5'].value = 'Territory Manager'
            else:
                sheet['B5'].value = 'Account Manager'
            sheet['B6'].value = comp_month
            wb.save(excel_file)
            wb.close()
            # break

            # set kwarg export=True to generate a PDF statement from the Excel file
            export = kwargs.get('export', None)
            if export:
                export_to_pdf("AMExportPDF")
        else:
            continue


def rm_statement(payees, **kwargs):
    """Exports all comp details to COMP_STATEMENT_RM.xlsx, which can then be used to generate an official statement.
        Optionally, you can pass in an email kwarg to view info for a single rep. Set export=True to generate a PDF
        statement."""
    email = kwargs.get('email', None)
    print("Generating RM Statements...")
    for rm in payees.rm_info:
        if email is None or payees.rm_info[rm]['EMAIL'] in email:
            # get the info for only the current loop rep
            name = rm
            eid = payees.rm_info[rm]['EMAIL']
            region = payees.rm_info[rm]['REGION']
            payout_df = payees.tblpayout[payees.tblpayout['EID'] == eid]
            comp_detail_df = payees.rm_comp_detail[payees.rm_comp_detail['SALES_CREDIT_RM_EMAIL'] == eid]

            excel_file = rm_comp_file

            # export the rep's tblPayout info and comp detail info to different tabs on COMP_STATEMENT.xlsx
            export_to_excel(excel_file, 'payout', payout_df)
            export_to_excel(excel_file, 'detail', comp_detail_df)

            wb = load_workbook(excel_file)
            sheet_name = 'info'
            sheet = wb[sheet_name]
            sheet['B1'].value = name
            sheet['B2'].value = eid
            sheet['B4'].value = region
            if payees.rm_info[rm]['EMAIL'] in rms:
                sheet['B5'].value = 'Region Manager'
            else:
                sheet['B5'].value = 'Area Director'
            sheet['B6'].value = comp_month
            wb.save(excel_file)
            wb.close()

            # set kwarg export=True to generate a PDF statement from the Excel file
            export = kwargs.get('export', None)
            if export:
                export_to_pdf("RMExportPDF")
        else:
            continue


def csr_statement(payees, **kwargs):
    """Exports all comp details to COMP_STATEMENT_CSR.xlsx, which can then be used to generate an official statement.
        Optionally, you can pass in an email kwarg to view info for a single rep. Set export=True to generate a PDF
        statement."""
    email = kwargs.get('email', None)
    print("Generating CSR Statements...")
    for csr in payees.csr_info:
        if email is None or payees.csr_info[csr]['EMAIL'] in email:
            # get the info for only the current loop rep
            name = csr
            eid = payees.csr_info[csr]['EMAIL']
            rm = payees.csr_info[csr]['RM_EMAIL']
            terr = payees.csr_info[csr]['TERR_NM']
            base_bonus = payees.csr_info[csr]['BASE_BONUS']
            quota = payees.csr_info[csr]['QUOTA']
            payout_df = payees.tblpayout[payees.tblpayout['EID'] == eid]
            comp_detail_df = payees.csr_comp_detail[payees.csr_comp_detail['SALES_CREDIT_FCE_EMAIL'] == eid]

            excel_file = csr_comp_file

            # export the rep's tblPayout info and comp detail info to different tabs on COMP_STATEMENT.xlsx
            export_to_excel(excel_file, 'payout', payout_df)
            export_to_excel(excel_file, 'detail', comp_detail_df)

            wb = load_workbook(excel_file)
            sheet_name = 'info'
            sheet = wb[sheet_name]
            sheet['B1'].value = name
            sheet['B2'].value = eid
            sheet['B3'].value = rm
            sheet['B4'].value = terr
            sheet['B6'].value = base_bonus
            sheet['B7'].value = comp_month
            sheet['B8'].value = quota
            wb.save(excel_file)
            wb.close()

            # set kwarg export=True to generate a PDF statement from the Excel file
            export = kwargs.get('export', None)
            if export:
                export_to_pdf("CSRExportPDF")
        else:
            continue
